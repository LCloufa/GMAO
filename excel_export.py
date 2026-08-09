from io import BytesIO
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


TITLE_FILL = PatternFill("solid", fgColor="2563EB")
CLIENT_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FILL = PatternFill("solid", fgColor="DCE6F1")
ALT_FILL = PatternFill("solid", fgColor="F8FAFC")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=16)
CLIENT_FONT = Font(color="FFFFFF", bold=True, size=12)
HEADER_FONT = Font(color="1F2937", bold=True)
THIN = Side(style="thin", color="D1D5DB")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def _sheet(wb, candidates, canonical_name):
    lower_map = {name.lower(): name for name in wb.sheetnames}
    for candidate in candidates:
        actual = lower_map.get(candidate.lower())
        if actual:
            return wb[actual]
    return wb.create_sheet(canonical_name)


def _reset_sheet(ws):
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)
    for name in list(ws.tables.keys()):
        del ws.tables[name]
    ws.freeze_panes = None
    ws.auto_filter.ref = None
    ws.sheet_view.showGridLines = False


def _apply_title(ws, end_col, title, subtitle):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    title_cell = ws.cell(1, 1, title)
    title_cell.fill = TITLE_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = LEFT
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    subtitle_cell = ws.cell(2, 1, subtitle)
    subtitle_cell.font = Font(color="64748B", italic=True, size=10)
    subtitle_cell.alignment = LEFT
    ws.row_dimensions[2].height = 20


def _write_headers(ws, row, headers):
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row, col, label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER
    ws.row_dimensions[row].height = 22


def _write_equipment_tree(ws, clients, equip_by_client, unassigned):
    _reset_sheet(ws)
    _apply_title(
        ws,
        8,
        "LISTING ÉQUIPEMENTS PAR CLIENT",
        f"Export généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
    )

    headers = [
        "Équipement",
        "Code",
        "Type",
        "Emplacement",
        "N° série",
        "Fabricant",
        "Modèle",
        "Statut",
    ]
    widths = [34, 18, 24, 24, 22, 24, 24, 22]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.freeze_panes = "A4"

    row = 4
    groups = [(cid, name, equip_by_client.get(cid, [])) for cid, name in clients]
    if unassigned:
        groups.append((None, "SANS CLIENT", unassigned))

    if not groups:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row, 1, "Aucun client ni équipement à exporter.")
        cell.alignment = LEFT
        cell.font = Font(italic=True, color="64748B")
        return

    for _, client_name, equipment_rows in groups:
        client_row = row
        ws.merge_cells(start_row=client_row, start_column=1, end_row=client_row, end_column=8)
        client_cell = ws.cell(
            client_row,
            1,
            f"{client_name}  —  {len(equipment_rows)} équipement(s)",
        )
        client_cell.fill = CLIENT_FILL
        client_cell.font = CLIENT_FONT
        client_cell.alignment = LEFT
        ws.row_dimensions[client_row].height = 24
        row += 1

        header_row = row
        _write_headers(ws, header_row, headers)
        row += 1

        if equipment_rows:
            for line_index, eq in enumerate(equipment_rows):
                values = [
                    eq.get("nom"),
                    eq.get("code"),
                    eq.get("type"),
                    eq.get("emplacement"),
                    eq.get("numero_serie"),
                    eq.get("fabricant"),
                    eq.get("modele"),
                    eq.get("statut"),
                ]
                for col, value in enumerate(values, start=1):
                    cell = ws.cell(row, col, value if value not in (None, "") else "-")
                    cell.border = THIN_BORDER
                    cell.alignment = LEFT if col in {1, 3, 4, 6, 7, 8} else CENTER
                    if line_index % 2:
                        cell.fill = ALT_FILL
                row += 1
        else:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            cell = ws.cell(row, 1, "Aucun équipement rattaché à ce client")
            cell.font = Font(italic=True, color="64748B")
            cell.alignment = LEFT
            for col in range(1, 9):
                ws.cell(row, col).border = THIN_BORDER
            row += 1

        body_end = row - 1
        ws.row_dimensions.group(header_row, body_end, outline_level=1, hidden=False)
        row += 1

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def _write_hours(ws, clients, hours_by_client):
    _reset_sheet(ws)
    _apply_title(
        ws,
        2,
        "HEURES D'INTERVENTION PAR CLIENT",
        "Somme des durées estimées des interventions enregistrées dans la GMAO",
    )
    _write_headers(ws, 4, ["Client", "Heures"])
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 16
    ws.freeze_panes = "A5"

    row = 5
    for index, (client_id, client_name) in enumerate(clients):
        values = [client_name, hours_by_client.get(client_id, 0.0)]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row, col, value)
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col == 1 else CENTER
            if index % 2:
                cell.fill = ALT_FILL
        ws.cell(row, 2).number_format = "0.00"
        row += 1

    if not clients:
        ws.merge_cells("A5:B5")
        ws["A5"] = "Aucun client à exporter."
        ws["A5"].font = Font(italic=True, color="64748B")


def _write_interventions(ws, interventions):
    _reset_sheet(ws)
    _apply_title(
        ws,
        6,
        "LISTING DES INTERVENTIONS",
        "Interventions enregistrées dans la GMAO",
    )
    headers = [
        "Titre intervention",
        "Équipement",
        "Technicien",
        "Type",
        "Date prévue",
        "Durée estimée (min)",
    ]
    _write_headers(ws, 4, headers)
    widths = [38, 30, 18, 18, 22, 22]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "A5"

    row = 5
    for index, item in enumerate(interventions):
        date_prevue = item.get("scheduled_date") or "-"
        if item.get("scheduled_time"):
            date_prevue = f"{date_prevue} {item['scheduled_time']}"
        values = [
            item.get("title") or "-",
            item.get("equipment") or "-",
            item.get("technician") or "-",
            item.get("type") or "-",
            date_prevue,
            item.get("estimated_duration") if item.get("estimated_duration") is not None else "-",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row, col, value)
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col in {1, 2} else CENTER
            if index % 2:
                cell.fill = ALT_FILL
        row += 1

    if not interventions:
        ws.merge_cells("A5:F5")
        ws["A5"] = "Aucune intervention à exporter."
        ws["A5"].font = Font(italic=True, color="64748B")

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def create_gmao_excel_export(get_db_connection, template_path):
    template = Path(template_path)
    if not template.is_file():
        raise FileNotFoundError(f"Modèle Excel introuvable : {template}")

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT id, nom FROM clients")
    clients = [(row[0], row[1]) for row in cursor.fetchall()]
    clients.sort(key=lambda item: ((item[1] or "").casefold(), item[0]))

    cursor.execute(
        """
        SELECT client_id, nom, code, type, emplacement, numero_serie,
               fabricant, modele, statut
        FROM equipements
        """
    )
    equip_by_client = {}
    unassigned = []
    for row in cursor.fetchall():
        item = {
            "client_id": row[0],
            "nom": row[1],
            "code": row[2],
            "type": row[3],
            "emplacement": row[4],
            "numero_serie": row[5],
            "fabricant": row[6],
            "modele": row[7],
            "statut": row[8],
        }
        if item["client_id"] is None:
            unassigned.append(item)
        else:
            equip_by_client.setdefault(item["client_id"], []).append(item)

    for rows in equip_by_client.values():
        rows.sort(key=lambda item: (item.get("nom") or "").casefold())
    unassigned.sort(key=lambda item: (item.get("nom") or "").casefold())

    cursor.execute(
        """
        SELECT c.id,
               COALESCE(SUM(COALESCE(i.estimated_duration, 0)), 0)
        FROM clients c
        LEFT JOIN equipements e ON e.client_id = c.id
        LEFT JOIN interventions i ON i.equipment_id = e.id
        GROUP BY c.id
        """
    )
    hours_by_client = {
        row[0]: round(float(row[1] or 0) / 60.0, 2)
        for row in cursor.fetchall()
    }

    cursor.execute(
        """
        SELECT i.title,
               e.nom,
               COALESCE(t.code, '-'),
               i.type,
               i.scheduled_date,
               i.scheduled_time,
               i.estimated_duration
        FROM interventions i
        LEFT JOIN equipements e ON e.id = i.equipment_id
        LEFT JOIN techniciens t ON t.id = i.assigned_to
        """
    )
    interventions = [
        {
            "title": row[0],
            "equipment": row[1],
            "technician": row[2],
            "type": row[3],
            "scheduled_date": row[4],
            "scheduled_time": row[5],
            "estimated_duration": row[6],
        }
        for row in cursor.fetchall()
    ]
    interventions.sort(
        key=lambda item: (
            str(item.get("scheduled_date") or ""),
            str(item.get("scheduled_time") or ""),
            str(item.get("title") or "").casefold(),
        ),
        reverse=True,
    )

    conn.close()

    wb = load_workbook(template)
    ws_eq = _sheet(
        wb,
        ["Listing équip", "Listing equip", "Listing équipement", "Listing equipement"],
        "Listing équip",
    )
    ws_h = _sheet(wb, ["Listing heures", "Listing heure"], "Listing heures")
    ws_i = _sheet(
        wb,
        ["Listing inter", "Listing intervention", "Listing interventions"],
        "Listing inter",
    )

    _write_equipment_tree(ws_eq, clients, equip_by_client, unassigned)
    _write_hours(ws_h, clients, hours_by_client)
    _write_interventions(ws_i, interventions)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
